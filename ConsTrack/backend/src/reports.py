import argparse
import io
import json
import math
import os
import sys
import random
from reportlab.lib.pagesizes import A4
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    HRFlowable, Image, KeepTogether
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.platypus import Flowable
from reportlab.lib.utils import ImageReader  

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np

LIGHT_BG  = "FFF0F4F8"
NAVY_HEX  = "0D1B2A"
STEEL_HEX = "1B3A5C"
ACCENT_HEX = "00C6A7"
ACCENT2   = "F4A261"
MID_GRAY  = "8A9BB0"
GREEN_OK  = "06D6A0"
RED_ZONE  = "E63946"
WHITE     = "FFFFFFFF"
YELLOW_H  = "FFD166"

def rl_color(hex_value):
    hex_value = str(hex_value).lstrip('#')
    if len(hex_value) == 8:
        hex_value = hex_value[2:]
    return colors.HexColor(f'#{hex_value}')

W, H = A4

# משתנים גלובליים שיאותחלו ב-main
ZONES = []
RUN_ID = ""
GENERATED = ""
T1_SCAN = ""
T2_SCAN = ""
OVERALL_PCT = 0.0
VOL_T1 = 0.0
VOL_T2 = 0.0
VOL_DELTA = 0.0
FORECAST = ""

def fig_to_image(fig, dpi=150):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', facecolor='none', transparent=True)
    buf.seek(0)
    plt.close(fig)
    return buf





def mpl_style():
    plt.rcParams.update({
        'font.family': 'DejaVu Sans',
        'axes.facecolor': '#FFF0F4F8',
        'figure.facecolor': 'none',
        'axes.edgecolor': '#8A9BB0',
        'axes.labelcolor': '#0D1B2A',
        'xtick.color': '#0D1B2A',
        'ytick.color': '#0D1B2A',
        'grid.color': '#FFFFFF',
        'grid.linewidth': 0.8,
        'axes.grid': True,
        'axes.spines.top': False,
        'axes.spines.right': False,
    })

# גרף חדש ואמיתי: מעקב נפח בין סריקות T1 ל-T2
def chart_volume_trend():
    mpl_style()
    fig, ax = plt.subplots(figsize=(7, 3.2))
    
    scans = [f"T1 Scan\n({T1_SCAN})", f"T2 Scan\n({T2_SCAN})"]
    volumes = [VOL_T1, VOL_T2]
    
    # ציור קו המגמה בין הנקודות האמיתיות
    ax.plot(scans, volumes, '-', color='#00C6A7', linewidth=3, marker='o', markersize=8, label='Measured Volume')
    ax.fill_between(scans, volumes, alpha=0.15, color='#00C6A7')
    
    # הוספת ערכי הנפח מעל הנקודות
    for i, vol in enumerate(volumes):
        ax.text(i, vol + (max(volumes) * 0.02), f"{vol:.3f} m³", ha='center', va='bottom', fontweight='bold', color='#0D1B2A', fontsize=10)
        
    ax.set_ylabel('Volume (m³)', fontsize=10)
    ax.set_title('Volume Change Trend (T1 to T2)', fontsize=12, fontweight='bold', color='#0D1B2A', pad=12)
    
    # התאמת גבולות ה-Y שיראה פרופורציונלי
    margin = max(max(volumes) * 0.1, 1.0)
    ax.set_ylim(max(0, min(volumes) - margin), max(volumes) + margin)
    
    fig.tight_layout()
    return fig_to_image(fig)

def chart_zone_progress():
    mpl_style()
    fig, ax = plt.subplots(figsize=(7, 3.2))
    names = [z['name'] for z in ZONES]
    progress = [z['progress'] for z in ZONES]
    completion = [z['completion'] for z in ZONES]
    y = np.arange(len(names))
    h = 0.35
    colors_prog = ['#00C6A7' if p >= c else '#E63946' for p, c in zip(progress, completion)]
    ax.barh(y + h/2, progress, h, color=colors_prog, alpha=0.85, label='Current Progress', zorder=3)
    ax.barh(y - h/2, completion, h, color='#1B3A5C', alpha=0.6, label='Historical Target', zorder=3)
    for i, (p, c) in enumerate(zip(progress, completion)):
        ax.text(p + 0.5, i + h/2, f'{p:.1f}%', va='center', fontsize=8, color='#0D1B2A')
    ax.set_yticks(y)
    ax.set_yticklabels(names, fontsize=9)
    ax.set_xlabel('Completion (%)', fontsize=9)
    ax.set_xlim(0, 110)
    ax.legend(fontsize=9, loc='lower right')
    ax.set_title('Zone-by-Zone Progress vs. Target', fontsize=11, fontweight='bold', color='#0D1B2A', pad=10)
    fig.tight_layout()
    return fig_to_image(fig)

def chart_donut():
    fig, ax = plt.subplots(figsize=(3.2, 3.2), facecolor='none')
    done = OVERALL_PCT
    left = 100 - done
    wedge_colors = ['#00C6A7', '#E8EEF4']
    ax.pie([done, left], colors=wedge_colors, startangle=90, wedgeprops=dict(width=0.45, edgecolor='white', linewidth=2))
    ax.text(0, 0, f'{done:.1f}%', ha='center', va='center', fontsize=22, fontweight='bold', color='#0D1B2A')
    ax.text(0, -0.22, 'OVERALL', ha='center', va='center', fontsize=8, color='#8A9BB0', fontweight='bold')
    fig.tight_layout()
    return fig_to_image(fig, dpi=180)

class SectionHeader(Flowable):
    def __init__(self, number, title, page_width):
        super().__init__()
        self.number = number
        self.title  = title
        self.width  = page_width
    def wrap(self, avail_w, avail_h):
        return self.width, 28
    def draw(self):
        c = self.canv
        c.setFillColor(colors.navy)
        c.rect(0, 0, self.width, 28, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.rect(0, 0, 6, 28, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('Helvetica-Bold', 12)
        c.drawString(16, 9, f'{self.number}.  {self.title}')


def build_pdf(target_pdf_path):
    doc = SimpleDocTemplate(target_pdf_path, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=14*mm, bottomMargin=14*mm)
    
    body = ParagraphStyle('body', fontName='Helvetica', fontSize=9.5, leading=14, textColor=rl_color(NAVY_HEX), spaceAfter=4)
    small = ParagraphStyle('small', fontName='Helvetica', fontSize=8.5, leading=12, textColor=rl_color(MID_GRAY))
    kv_key = ParagraphStyle('kk', fontName='Helvetica-Bold', fontSize=9, textColor=rl_color(STEEL_HEX))
    kv_val = ParagraphStyle('kv', fontName='Helvetica', fontSize=9, textColor=rl_color(NAVY_HEX))
    th_style = ParagraphStyle('th', fontName='Helvetica-Bold', fontSize=9, textColor=rl_color(WHITE))
    
    content_width = W - 36*mm
    story = []

    class CoverBanner(Flowable):
        def wrap(self, aw, ah): return content_width, 110
        def draw(self):
            c = self.canv
            c.setFillColor(rl_color(NAVY_HEX))
            c.roundRect(0, 0, content_width, 110, 8, fill=1, stroke=0)
            c.setFillColor(rl_color(ACCENT_HEX))
            c.rect(0, 0, content_width, 5, fill=1, stroke=0)
            buf = chart_donut()
            img = ImageReader(buf)
            c.drawImage(img, 12, 15, width=80, height=80, mask='auto')
            c.setFillColor(rl_color(ACCENT_HEX))
            c.setFont('Helvetica-Bold', 28)
            c.drawString(108, 68, 'CONSTRACK')
            c.setFillColor(colors.white)
            c.setFont('Helvetica-Bold', 13)
            c.drawString(108, 48, 'PROGRESS REPORT')
            c.setFillColor(colors.HexColor("#8A9BB0"))
            c.setFont('Helvetica', 8.5)
            c.drawString(108, 33, f'Generated: {GENERATED[:10]}')
            c.drawString(108, 21, f'Run ID: {RUN_ID[:24]}...')

    story.append(CoverBanner())
    story.append(Spacer(1, 10))

    meta_rows = [['Run ID', RUN_ID], ['T1 Scan Date', T1_SCAN], ['T2 Scan Date', T2_SCAN], ['Forecast', FORECAST]]
    meta_table_data = [[Paragraph(k, kv_key), Paragraph(v, kv_val)] for k, v in meta_rows]
    meta_tbl = Table(meta_table_data, colWidths=[content_width*0.28, content_width*0.72])
    meta_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), rl_color(LIGHT_BG)),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [rl_color(LIGHT_BG), rl_color(WHITE)]),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
        ('BOX', (0,0), (-1,-1), 0.5, rl_color(MID_GRAY)),
        ('LINEBELOW', (0,0), (-1,-2), 0.3, colors.HexColor('#D0D8E4')),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 14))

    story.append(SectionHeader('1', 'Progress by Area', content_width))
    story.append(Spacer(1, 8))

    cw = content_width / 3 - 3
    kpi_tbl = Table([[
        Table([[Paragraph(f'<font color="#00C6A7"><b>{OVERALL_PCT:.1f}%</b></font>', ParagraphStyle('k1', fontName='Helvetica-Bold', fontSize=22, textColor=rl_color(ACCENT_HEX), leading=26))], [Paragraph('Overall Progress', small)]], colWidths=[cw]),
        Table([[Paragraph(f'<font color="#F4A261"><b>{FORECAST}</b></font>', ParagraphStyle('k2', fontName='Helvetica-Bold', fontSize=16, textColor=rl_color(ACCENT2), leading=20))], [Paragraph('Forecast Completion', small)]], colWidths=[cw]),
        Table([[Paragraph(f'<font color="#06D6A0"><b>{VOL_DELTA:+.3f} m³</b></font>', ParagraphStyle('k3', fontName='Helvetica-Bold', fontSize=18, textColor=rl_color(GREEN_OK), leading=22))], [Paragraph('Net Volume Change', small)]], colWidths=[cw]),
    ]], colWidths=[cw+3]*3)
    kpi_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), rl_color(LIGHT_BG)),
        ('BOX', (0,0), (-1,-1), 0.4, rl_color(MID_GRAY)),
        ('INNERGRID', (0,0), (-1,-1), 0.4, colors.HexColor('#D0D8E4')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(kpi_tbl)
    story.append(Spacer(1, 10))

    if ZONES:
        buf_zone = chart_zone_progress()
        story.append(Image(buf_zone, width=content_width, height=content_width * 0.46))
        story.append(Spacer(1, 8))

        zone_header = [Paragraph('<b>Zone</b>', th_style), Paragraph('<b>Type</b>', th_style), Paragraph('<b>Target %</b>', th_style), Paragraph('<b>Progress %</b>', th_style), Paragraph('<b>DeltaVol (m3)</b>', th_style)]
        zone_rows = [zone_header]
        for z in ZONES:
            zone_rows.append([
                Paragraph(z['name'], body), Paragraph(z['type'], body),
                Paragraph(f"{z['completion']:.1f}%", body), Paragraph(f"<b>{z['progress']:.1f}%</b>", body),
                Paragraph(f"{z['vol']:.3f}", body),
            ])
        zone_tbl = Table(zone_rows, colWidths=[content_width*w for w in [0.25, 0.16, 0.14, 0.17, 0.16]])
        zone_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), rl_color(NAVY_HEX)),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [rl_color(WHITE), rl_color(LIGHT_BG)]),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'), ('FONTSIZE', (0,0), (-1,-1), 8.5),
            ('TOPPADDING', (0,0), (-1,-1), 5), ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 6), ('BOX', (0,0), (-1,-1), 0.5, rl_color(MID_GRAY)),
            ('LINEBELOW', (0,0), (-1,-1), 0.3, colors.HexColor('#D0D8E4')),
            ('ALIGN', (2,0), (-1,-1), 'CENTER'),
        ]))
        story.append(zone_tbl)
        story.append(Spacer(1, 14))

    story.append(SectionHeader('2', 'Volume & Area Material Changes', content_width))
    story.append(Spacer(1, 8))

    vol_data = [
        [Paragraph('<b>Metric</b>', th_style), Paragraph('<b>Value</b>', th_style)],
        [Paragraph('Initial Volume - T1 Scan', body), Paragraph(f'{VOL_T1:.3f} m3', body)],
        [Paragraph('Current Volume - T2 Scan', body), Paragraph(f'{VOL_T2:.3f} m3', body)],
        [Paragraph('Net Volume Change (DeltaV)', body), Paragraph(f'<font color="#E63946"><b>{VOL_DELTA:.3f} m3</b></font>', body)],
    ]
    vol_tbl = Table(vol_data, colWidths=[content_width*0.55, content_width*0.45])
    vol_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), rl_color(STEEL_HEX)),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [rl_color(WHITE), rl_color(LIGHT_BG)]),
        ('BOX', (0,0), (-1,-1), 0.5, rl_color(MID_GRAY)),
        ('LINEBELOW', (0,0), (-1,-1), 0.3, colors.HexColor('#D0D8E4')),
        ('TOPPADDING', (0,0), (-1,-1), 6), ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(vol_tbl)
    story.append(Spacer(1, 14))

    story.append(SectionHeader('3', 'Volume Analysis Graph', content_width))
    story.append(Spacer(1, 8))
    
    # הזרקת הגרף החדש המבוסס על הנקודות האמיתיות
    buf_trend = chart_volume_trend()
    story.append(Image(buf_trend, width=content_width, height=content_width * 0.46))
    story.append(Spacer(1, 10))

    story.append(SectionHeader('4', 'Forecast Completion Estimation', content_width))
    story.append(Spacer(1, 8))

    
    


    forecast_text = (f'Based on current site velocity and cross-scan comparative comparison, '
                     f'the total registered change in site material is <b>{VOL_DELTA:+.3f} m³</b>. '
                     f'The project trajectory is monitored dynamically with target forecast completion estimated for '
                     f'<font color="#F4A261"><b>{FORECAST}</b></font>.')
    story.append(Paragraph(forecast_text, body))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width=content_width, thickness=1, color=rl_color(ACCENT_HEX), spaceAfter=6))
    story.append(Paragraph(f'CONSTRACK Automated Report  -  Run {RUN_ID[:16]}...  -  {GENERATED[:10]}  -  Confidential.',
        ParagraphStyle('footer', fontName='Helvetica', fontSize=7.5, textColor=rl_color(MID_GRAY))))

    doc.build(story)


from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def style_cell(ws, row, col, value=None, bold=False, font_color="000000",
               bg=None, align='left', size=10, border=True, number_format=None):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = Font(name='Arial', bold=bold, color=font_color, size=size)
    if bg:
        cell.fill = PatternFill('solid', fgColor=bg)
    if border:
        thin = Side(style='thin', color='D0D8E4')
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if number_format:
        cell.number_format = number_format
    return cell

def merge_title(ws, row, col_start, col_end, text, bg=NAVY_HEX, fg=WHITE, sz=12):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start)
    cell.value = text
    cell.font = Font(name='Arial', bold=True, color=fg, size=sz)
    cell.fill = PatternFill('solid', fgColor=bg)

def build_excel(target_xlsx_path):
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Overview")
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 26
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 18
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 18
    ws1.row_dimensions[1].height = 40

    ws1.merge_cells('A1:F1')
    ws1['A1'].value = '  CONSTRACK — PROGRESS REPORT'
    ws1['A1'].font = Font(name='Arial', bold=True, color=ACCENT_HEX, size=22)
    ws1['A1'].fill = PatternFill('solid', fgColor=NAVY_HEX)

    ws1.merge_cells('A3:F3')
    ws1['A3'].value = f'  Generated: {GENERATED}   |   Run ID: {RUN_ID}   |   Forecast: {FORECAST}'
    ws1['A3'].font = Font(name='Arial', color=MID_GRAY, size=9)
    ws1['A3'].fill = PatternFill('solid', fgColor="E8EEF4")

    kpis = [
        ('Overall Progress', f'{OVERALL_PCT:.1f}%', ACCENT_HEX),
        ('Active Zones', str(len(ZONES)), STEEL_HEX),
        ('Est. Completion', FORECAST, ACCENT2),
    ]
    for i, (label, value, color) in enumerate(kpis, start=3):
        col = i
        style_cell(ws1, 6, col, value=value, bold=True, font_color=color, bg=LIGHT_BG, align='center', size=16, border=False)
        style_cell(ws1, 7, col, value=label, bold=False, font_color=MID_GRAY, bg=LIGHT_BG, align='center', size=9, border=False)
        ws1.cell(row=6, column=col).border = Border(bottom=Side(style='medium', color=color))

    merge_title(ws1, 10, 1, 6, '  Run Metadata', bg=STEEL_HEX)
    meta = [('Run ID', RUN_ID), ('T1 Scan Date', T1_SCAN), ('T2 Scan Date', T2_SCAN), ('Forecast', FORECAST), ('Generated', GENERATED)]
    for i, (k, v) in enumerate(meta, start=11):
        style_cell(ws1, i, 1, value=k, bold=True, font_color=STEEL_HEX, bg=LIGHT_BG if i%2==0 else WHITE, size=9)
        style_cell(ws1, i, 2, value=v, bold=False, font_color=NAVY_HEX, bg=LIGHT_BG if i%2==0 else WHITE, size=9)

    merge_title(ws1, 19, 1, 6, '  Volume Summary', bg=STEEL_HEX)
    vol_data = [('Initial Volume (T1 Scan)', VOL_T1, 'm³'), ('Current Volume (T2 Scan)', VOL_T2, 'm³'), ('Net Volume Change (ΔV)', VOL_DELTA, 'm³')]
    for i, (label, val, unit) in enumerate(vol_data, start=20):
        bg = LIGHT_BG if i%2==0 else WHITE
        style_cell(ws1, i, 1, value=label, bold=False, font_color=NAVY_HEX, bg=bg, size=9)
        style_cell(ws1, i, 2, value=val, bold=True, font_color=RED_ZONE if val < 0 else NAVY_HEX, bg=bg, size=10, number_format='#,##0.000')
        style_cell(ws1, i, 3, value=unit, bold=False, font_color=MID_GRAY, bg=bg, size=9)

    if ZONES:
        ws2 = wb.create_sheet("Zone Progress")
        ws2.sheet_view.showGridLines = False
        for col, w in zip(['A','B','C','D','E'], [20,14,14,14,16]): ws2.column_dimensions[col].width = w
        ws2.merge_cells('A1:G1')
        ws2['A1'].value = '  ZONE PROGRESS & STATUS'
        ws2['A1'].font = Font(name='Arial', bold=True, color=ACCENT_HEX, size=16)
        ws2['A1'].fill = PatternFill('solid', fgColor=NAVY_HEX)
        
        headers = ['Zone', 'Type', 'Target %', 'Progress %', 'ΔVol (m³)']
        for col, h in enumerate(headers, 1): style_cell(ws2, 3, col, value=h, bold=True, font_color=WHITE, bg=NAVY_HEX, align='center')
        
        for i, z in enumerate(ZONES, start=4):
            bg = LIGHT_BG if i%2==0 else WHITE
            style_cell(ws2, i, 1, value=z['name'], bold=True, font_color=NAVY_HEX, bg=bg)
            style_cell(ws2, i, 2, value=z['type'], font_color=MID_GRAY, bg=bg, align='center')
            style_cell(ws2, i, 3, value=z['completion']/100, font_color=STEEL_HEX, bg=bg, align='center', number_format='0.0%')
            style_cell(ws2, i, 4, value=z['progress']/100, bold=True, font_color=ACCENT_HEX, bg=bg, align='center', number_format='0.0%')
            style_cell(ws2, i, 5, value=z['vol'], font_color=RED_ZONE, bg=bg, align='center', number_format='#,##0.000')

    ws1.freeze_panes = 'A4'
    ws1.sheet_properties.tabColor = ACCENT_HEX
    wb.save(target_xlsx_path)

def main():
    global RUN_ID, GENERATED, T1_SCAN, T2_SCAN, OVERALL_PCT, VOL_T1, VOL_T2, VOL_DELTA, FORECAST, ZONES

    if len(sys.argv) < 2:
        print("[DEBUG Python ERROR] No arguments passed to script!", flush=True)
        sys.exit(1)

    try:
        print("[DEBUG Python] Attempting to parse JSON input...", flush=True)
        input_data = json.loads(sys.argv[1])
        print("[DEBUG Python] JSON parsed successfully!", flush=True)
        
        out_dir = input_data["outDir"]
        pdf_path = input_data["pdfPath"]
        xlsx_path = input_data["xlsxPath"]
        
        run_obj = input_data.get("run", {})
        zones_list = input_data.get("zones", [])

        RUN_ID = str(run_obj.get("_id", "Unknown_Run"))
        GENERATED = str(run_obj.get("createdAtISO", ""))
        
        # שמירת תאריך נקי (10 תווים ראשונים של ה-ISO String)
        T1_SCAN = str(input_data.get("t1ScanDate", "Unknown_T1"))[:10] 
        T2_SCAN = str(input_data.get("t2ScanDate", "Unknown_T2"))[:10] 
        
        print(f"[DEBUG Python Scans Match] Parsed values -> T1 Scan Date: '{T1_SCAN}' | T2 Scan Date: '{T2_SCAN}'", flush=True)

        OVERALL_PCT = float(run_obj.get("overallProgressPct", 0.0))
        VOL_T1 = float(run_obj.get("volumeT1M3", 0.0))
        VOL_T2 = float(run_obj.get("volumeT2M3", 0.0))
        VOL_DELTA = float(run_obj.get("volumeChangeM3", 0.0))
        FORECAST = str(run_obj.get("etaISO", ""))[:10]  

        raw_forecast = run_obj.get("etaISO")
        print(f"[DEBUG Forecast] Raw value from Node: '{raw_forecast}'", flush=True)

        FORECAST = str(raw_forecast or "")[:10]
        if not FORECAST:
            print("[DEBUG Forecast WARNING] FORECAST parsed as empty string!", flush=True)

        ZONES = []
        metrics_by_zone = {m["zoneId"]: m for m in run_obj.get("metricsByZone", [])}
        
        for z in zones_list:
            z_id = z.get("_id")
            metric = metrics_by_zone.get(z_id, {})
            ZONES.append({
                'name': z.get("name", "Unknown Zone"),
                'type': z.get("type", "site").capitalize(),
                'completion': float(z.get("completionPct", 0.0)),
                'progress': float(metric.get("progressPct", 0.0)),
                'vol': float(metric.get("volumeChangeM3", 0.0))
            })

        print(f"[DEBUG Python] Target PDF path: {pdf_path}", flush=True)
        print(f"[DEBUG Python] Target XLSX path: {xlsx_path}", flush=True)
        
        os.makedirs(out_dir, exist_ok=True)
        
        print("[DEBUG Python] Generating PDF report...", flush=True)
        build_pdf(pdf_path)
        
        print("[DEBUG Python] Generating Excel report...", flush=True)
        build_excel(xlsx_path)
        
        print("[DEBUG Python] Real styled reports created successfully!", flush=True)
        print(json.dumps({"success": True}), flush=True)

    except Exception as e:
        print(f"[DEBUG Python ERROR] Exception occurred: {str(e)}", flush=True)
        sys.exit(1)

if __name__ == "__main__":
    main()